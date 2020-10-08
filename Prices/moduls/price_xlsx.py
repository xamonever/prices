import io
from openpyxl import load_workbook
from ..abstruct.abs_price import AbstractPrice
# from .components import *


class XLSXLoader(AbstractPrice):
    """docstring for XLSXLoader"""

    def __init__(self, instruction, ):
        self.f_format = 'xlsx'
        self.mainsheet = int(instruction['mainsheet']) if instruction['mainsheet'] else 0
        self.encoding = instruction['encoding']

    
    @staticmethod
    def get_headers(I_ws, content_start, column_order):
        if I_ws.max_row and I_ws.max_row <= content_start:
            I_ws.reset_dimensions()
            I_ws.calculate_dimension(force=True)
        return I_ws.iter_rows(min_col=min(column_order), max_col=max(column_order), max_row=content_start)
    
    @staticmethod
    def get_content(I_ws, content_start, column_order):
        if I_ws.max_row and I_ws.max_row <= content_start:
            I_ws.reset_dimensions()
            I_ws.calculate_dimension(force=True)
        return I_ws.iter_rows(min_col=min(column_order), max_col=max(column_order), min_row=content_start, max_row=I_ws.max_row)

    @staticmethod
    def get_row_vals(I_ws, row, column_order):
        return  [cell.value for cell in row]


    @staticmethod
    def init_attr_by_int_list(obj, attr, instr, delim=','):
        setattr(obj, attr, [int(i) for i in instr.split(delim)])


    def get_input_data(self, file_name):
        input_wb = self.open_wb(file_name)
        I_ws = input_wb.worksheets[self.mainsheet]
        return  (input_wb, I_ws)


    def open_wb(self, file_name):
        with open(file_name, "rb") as f:
            in_mem_file = io.BytesIO(f.read())
        return load_workbook(in_mem_file, read_only=True, data_only=True)


    def close_wb(self, book):
        book.close()
   